"""
Generador de reportes Excel para resultados del Agente 3.

Genera un Excel con 6 hojas:
1. Resumen Ejecutivo
2. Registros Exitosos
3. Registros Parciales
4. No Registradas
5. Conceptos Sin Match
6. Detalle de Matching
"""
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Optional

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.erp.models import ResultadoRegistro, ResultadoMatchProducto
from src.sat.models import Factura


# Estilos
FONT_TITULO = Font(name='Calibri', bold=True, size=14)
FONT_SUBTITULO = Font(name='Calibri', bold=True, size=11)
FONT_HEADER = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
FONT_NORMAL = Font(name='Calibri', size=10)
FONT_EXITO = Font(name='Calibri', size=10, color='006100')
FONT_ERROR = Font(name='Calibri', size=10, color='9C0006')

FILL_HEADER = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
FILL_EXITO = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_PARCIAL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
FILL_ERROR = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FILL_RESUMEN = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


class ExcelGenerator:
    """Generador de reportes Excel para Agente 3"""

    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generar(
        self,
        resultados: List[ResultadoRegistro],
        facturas: List[Factura],
        todos_matches: List[Tuple[Factura, List[ResultadoMatchProducto]]],
        dry_run: bool = False,
        duracion_segundos: float = 0,
    ) -> Path:
        """
        Generar reporte Excel completo.

        Args:
            resultados: Lista de ResultadoRegistro por cada factura
            facturas: Lista de Factura originales
            todos_matches: Tuplas (Factura, matches) para detalle
            dry_run: Si fue ejecucion en modo dry-run
            duracion_segundos: Duracion de la ejecucion en segundos

        Returns:
            Ruta del archivo Excel generado
        """
        wb = Workbook()

        # Hoja 1: Resumen Ejecutivo
        self._hoja_resumen(wb, resultados, dry_run, duracion_segundos)

        # Hoja 2: Registros Exitosos
        exitosos = [
            (r, f) for r, f in zip(resultados, facturas)
            if r.exito and not r.registro_parcial
        ]
        self._hoja_registros(wb, "Exitosos", exitosos, FILL_EXITO)

        # Hoja 3: Registros Parciales
        parciales = [
            (r, f) for r, f in zip(resultados, facturas)
            if r.exito and r.registro_parcial
        ]
        self._hoja_registros(wb, "Parciales", parciales, FILL_PARCIAL)

        # Hoja 4: No Registradas
        fallidos = [
            (r, f) for r, f in zip(resultados, facturas)
            if not r.exito
        ]
        self._hoja_no_registradas(wb, fallidos)

        # Hoja 5: Conceptos Sin Match
        self._hoja_conceptos_sin_match(wb, resultados, facturas)

        # Hoja 6: Detalle de Matching
        self._hoja_detalle_matching(wb, todos_matches)

        # Remover hoja default vacia si existe
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Guardar
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        prefijo = "DRYRUN_" if dry_run else ""
        nombre = f"{prefijo}registro_directo_{timestamp}.xlsx"
        ruta = self.output_dir / nombre

        wb.save(str(ruta))
        logger.info(f"Reporte guardado: {ruta}")
        return ruta

    def _hoja_resumen(
        self,
        wb: Workbook,
        resultados: List[ResultadoRegistro],
        dry_run: bool,
        duracion: float,
    ):
        """Hoja 1: Resumen Ejecutivo"""
        ws = wb.create_sheet("Resumen", 0)

        # Titulo
        ws.merge_cells('A1:F1')
        celda = ws['A1']
        celda.value = "Agente 3 - Registro Directo de Facturas CFDI"
        celda.font = FONT_TITULO
        celda.alignment = ALIGN_CENTER

        if dry_run:
            ws.merge_cells('A2:F2')
            celda = ws['A2']
            celda.value = "MODO DRY-RUN (SIN CAMBIOS EN BD)"
            celda.font = Font(name='Calibri', bold=True, size=12, color='FF0000')
            celda.alignment = ALIGN_CENTER

        # Info de ejecucion
        fila = 4
        info = [
            ("Fecha de ejecucion", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Duracion", f"{duracion:.1f} segundos"),
            ("Modo", "DRY-RUN (simulacion)" if dry_run else "PRODUCCION"),
        ]
        for label, valor in info:
            ws.cell(row=fila, column=1, value=label).font = FONT_SUBTITULO
            ws.cell(row=fila, column=2, value=valor).font = FONT_NORMAL
            fila += 1

        # Totales
        fila += 1
        total = len(resultados)
        exitosos = sum(1 for r in resultados if r.exito and not r.registro_parcial)
        parciales = sum(1 for r in resultados if r.exito and r.registro_parcial)
        fallidos = sum(1 for r in resultados if not r.exito)

        ws.cell(row=fila, column=1, value="TOTALES").font = FONT_SUBTITULO
        fila += 1

        metricas = [
            ("Facturas procesadas", total, None),
            ("Registros exitosos", exitosos, FILL_EXITO),
            ("Registros parciales", parciales, FILL_PARCIAL),
            ("No registradas", fallidos, FILL_ERROR),
        ]
        for label, valor, fill in metricas:
            c1 = ws.cell(row=fila, column=1, value=label)
            c2 = ws.cell(row=fila, column=2, value=valor)
            c1.font = FONT_NORMAL
            c2.font = FONT_SUBTITULO
            c2.alignment = ALIGN_CENTER
            if fill:
                c1.fill = fill
                c2.fill = fill
            fila += 1

        # Conceptos
        fila += 1
        total_conceptos = sum(r.total_conceptos for r in resultados)
        total_matcheados = sum(r.conceptos_matcheados_count for r in resultados)
        total_no_matcheados = total_conceptos - total_matcheados
        porc_match = (total_matcheados / total_conceptos * 100) if total_conceptos > 0 else 0

        ws.cell(row=fila, column=1, value="CONCEPTOS").font = FONT_SUBTITULO
        fila += 1
        for label, valor in [
            ("Total conceptos", total_conceptos),
            ("Matcheados", total_matcheados),
            ("Sin match", total_no_matcheados),
            ("Porcentaje match", f"{porc_match:.1f}%"),
        ]:
            ws.cell(row=fila, column=1, value=label).font = FONT_NORMAL
            ws.cell(row=fila, column=2, value=valor).font = FONT_NORMAL
            fila += 1

        # Errores por tipo
        errores_tipo = {}
        for r in resultados:
            if not r.exito and r.error:
                errores_tipo[r.error] = errores_tipo.get(r.error, 0) + 1

        if errores_tipo:
            fila += 1
            ws.cell(row=fila, column=1, value="ERRORES POR TIPO").font = FONT_SUBTITULO
            fila += 1
            for tipo, cuenta in sorted(errores_tipo.items(), key=lambda x: x[1], reverse=True):
                ws.cell(row=fila, column=1, value=tipo).font = FONT_NORMAL
                ws.cell(row=fila, column=2, value=cuenta).font = FONT_NORMAL
                fila += 1

        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25

    def _hoja_registros(
        self,
        wb: Workbook,
        nombre_hoja: str,
        registros: List[Tuple[ResultadoRegistro, Factura]],
        fill_fila: PatternFill,
    ):
        """Hojas 2 y 3: Registros exitosos / parciales"""
        ws = wb.create_sheet(nombre_hoja)

        headers = [
            "Factura ERP", "UUID", "RFC Emisor", "Nombre Emisor",
            "Folio XML", "Fecha Emision", "Subtotal", "IVA", "Total",
            "Conceptos Total", "Conceptos Match", "% Match", "Metodo Pago",
        ]
        self._escribir_headers(ws, headers)

        for fila_idx, (resultado, factura) in enumerate(registros, start=2):
            datos = [
                resultado.numero_factura_erp or '',
                resultado.factura_uuid,
                factura.rfc_emisor,
                factura.nombre_emisor,
                factura.folio or '',
                factura.fecha_emision.strftime('%Y-%m-%d') if factura.fecha_emision else '',
                float(factura.subtotal),
                float(factura.iva_trasladado),
                float(factura.total),
                resultado.total_conceptos,
                resultado.conceptos_matcheados_count,
                f"{resultado.porcentaje_matcheados:.0f}%",
                factura.metodo_pago.value if factura.metodo_pago else '',
            ]
            for col, valor in enumerate(datos, start=1):
                celda = ws.cell(row=fila_idx, column=col, value=valor)
                celda.font = FONT_NORMAL
                celda.border = BORDER_THIN

        self._ajustar_anchos(ws)

    def _hoja_no_registradas(
        self,
        wb: Workbook,
        registros: List[Tuple[ResultadoRegistro, Factura]],
    ):
        """Hoja 4: Facturas no registradas"""
        ws = wb.create_sheet("No Registradas")

        headers = [
            "UUID", "RFC Emisor", "Nombre Emisor", "Folio XML",
            "Fecha Emision", "Total", "Conceptos",
            "Error", "Mensaje",
        ]
        self._escribir_headers(ws, headers)

        for fila_idx, (resultado, factura) in enumerate(registros, start=2):
            datos = [
                resultado.factura_uuid,
                factura.rfc_emisor,
                factura.nombre_emisor,
                factura.folio or '',
                factura.fecha_emision.strftime('%Y-%m-%d') if factura.fecha_emision else '',
                float(factura.total),
                resultado.total_conceptos,
                resultado.error or '',
                resultado.mensaje,
            ]
            for col, valor in enumerate(datos, start=1):
                celda = ws.cell(row=fila_idx, column=col, value=valor)
                celda.font = FONT_NORMAL
                celda.border = BORDER_THIN
                if col >= 8:
                    celda.fill = FILL_ERROR

        self._ajustar_anchos(ws)

    def _hoja_conceptos_sin_match(
        self,
        wb: Workbook,
        resultados: List[ResultadoRegistro],
        facturas: List[Factura],
    ):
        """Hoja 5: Conceptos que no matchearon"""
        ws = wb.create_sheet("Sin Match")

        headers = [
            "UUID Factura", "Nombre Emisor", "Descripcion XML",
            "Clave SAT", "Cantidad", "Valor Unitario", "Importe",
            "Motivo",
        ]
        self._escribir_headers(ws, headers)

        fila = 2
        for resultado, factura in zip(resultados, facturas):
            for concepto_reg in resultado.conceptos_no_matcheados:
                concepto = concepto_reg.concepto_xml
                datos = [
                    resultado.factura_uuid[:12] + '...',
                    factura.nombre_emisor,
                    concepto.descripcion,
                    concepto.clave_prod_serv,
                    float(concepto.cantidad),
                    float(concepto.valor_unitario),
                    float(concepto.importe),
                    concepto_reg.motivo_no_registro or '',
                ]
                for col, valor in enumerate(datos, start=1):
                    celda = ws.cell(row=fila, column=col, value=valor)
                    celda.font = FONT_NORMAL
                    celda.border = BORDER_THIN
                fila += 1

        self._ajustar_anchos(ws)

    def _hoja_detalle_matching(
        self,
        wb: Workbook,
        todos_matches: List[Tuple[Factura, List[ResultadoMatchProducto]]],
    ):
        """Hoja 6: Detalle completo de matching"""
        ws = wb.create_sheet("Detalle Matching")

        headers = [
            "UUID", "Emisor", "Descripcion XML", "Clave SAT XML",
            "Matcheado", "Producto ERP", "Nombre ERP", "Metodo",
            "Confianza", "Nivel", "Candidatos Desc.", "Mensaje",
        ]
        self._escribir_headers(ws, headers)

        fila = 2
        for factura, matches in todos_matches:
            for match in matches:
                concepto = match.concepto_xml
                producto = match.producto_erp

                datos = [
                    factura.uuid[:12] + '...',
                    factura.nombre_emisor[:30],
                    concepto.descripcion,
                    concepto.clave_prod_serv,
                    "SI" if match.matcheado else "NO",
                    producto.codigo if producto else '',
                    producto.nombre if producto else '',
                    match.metodo_match,
                    f"{match.confianza:.0%}" if match.confianza > 0 else '',
                    match.nivel_confianza,
                    match.candidatos_descartados,
                    match.mensaje,
                ]
                for col, valor in enumerate(datos, start=1):
                    celda = ws.cell(row=fila, column=col, value=valor)
                    celda.font = FONT_NORMAL
                    celda.border = BORDER_THIN

                    # Color segun resultado
                    if col == 5:
                        if match.matcheado:
                            celda.fill = FILL_EXITO
                        else:
                            celda.fill = FILL_ERROR
                fila += 1

        self._ajustar_anchos(ws)

    def _escribir_headers(self, ws, headers: List[str]):
        """Escribir fila de headers con formato"""
        for col, header in enumerate(headers, start=1):
            celda = ws.cell(row=1, column=col, value=header)
            celda.font = FONT_HEADER
            celda.fill = FILL_HEADER
            celda.alignment = ALIGN_CENTER
            celda.border = BORDER_THIN

        # Filtros automaticos
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        # Congelar primera fila
        ws.freeze_panes = 'A2'

    def _ajustar_anchos(self, ws):
        """Ajustar ancho de columnas basado en contenido"""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)

            for cell in col_cells:
                if cell.value:
                    cell_len = len(str(cell.value))
                    max_len = max(max_len, cell_len)

            # Limitar ancho maximo
            ancho = min(max_len + 2, 50)
            ancho = max(ancho, 8)
            ws.column_dimensions[col_letter].width = ancho
